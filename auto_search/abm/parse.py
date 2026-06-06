"""Parse the multi-sheet ABM target workbook into TargetAccount rows.

The real file (Q2 accounts) varies column layout per sheet: the company-name
column is 'Hospital Name' on hospital sheets and 'Physician Group Name' on
group sheets; only some sheets carry Website / State. We detect columns by
header text, skip sheets with no name column (Filters, pivots), and expand
(FKA ...)/(AKA ...) aliases so a former name still matches.
"""

from __future__ import annotations

import io

from auto_search.abm.models import TargetAccount
from auto_search.abm.util import (
    bare_domain,
    extract_state,
    segment_for_sheet,
    split_aliases,
)
from auto_search.normalize import normalize_company_name


def _find_col(header: tuple, *needles: str) -> int | None:
    """Index of the first header cell containing any needle (case-insensitive).

    Needles are tried in order of preference, so a more specific header
    ('hospital name') wins over a generic one ('name') when both are passed.
    """
    cells = [str(h or "").strip().lower() for h in header]
    for needle in needles:
        for i, cell in enumerate(cells):
            if cell and needle in cell:
                return i
    return None


def _cell(row: tuple, idx: int | None) -> object:
    return row[idx] if idx is not None and idx < len(row) else None


def parse_workbook(data: bytes) -> list[TargetAccount]:
    """Parse every sheet of the workbook into a flat list of TargetAccount."""
    import openpyxl  # local: heavy import, only needed when a file is uploaded

    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    targets: list[TargetAccount] = []
    try:
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            rows = [
                r for r in ws.iter_rows(values_only=True)
                if any(c is not None and str(c).strip() for c in r)
            ]
            if not rows:
                continue
            header = rows[0]
            name_col = _find_col(header, "hospital name", "physician group name", "name")
            if name_col is None:           # not a company list (Filters, pivots)
                continue
            web_col = _find_col(header, "website", "url", "domain")
            state_col = _find_col(header, "state")
            id_col = _find_col(header, "definitive id")
            segment = segment_for_sheet(sheet)

            for row in rows[1:]:
                raw = _cell(row, name_col)
                raw = "" if raw is None else str(raw).strip()
                if not raw:
                    continue
                primary, aliases = split_aliases(raw)
                keys = sorted({
                    k for n in (primary, *aliases) if (k := normalize_company_name(n))
                })
                if not keys:
                    continue
                did = _cell(row, id_col)
                targets.append(TargetAccount(
                    name=primary or raw,
                    aliases=aliases,
                    keys=keys,
                    domain=bare_domain(_cell(row, web_col)),
                    state=extract_state(_cell(row, state_col)),
                    segment=segment,
                    source_sheet=sheet,
                    definitive_id=str(did).strip() if did not in (None, "") else None,
                ))
    finally:
        wb.close()
    return targets
