"""Styled Excel export of scored accounts — the human-facing counterpart of the
CSV export (which stays for data work). PURE: accounts in -> Workbook out, no
I/O; the endpoint streams it.

Two sheets:
  - "Scored accounts": one row per account. Fit band color-coded, frozen +
    filterable header, wrapped evidence/recommendation columns, banded rows.
  - "Summary": fit distribution, segment counts, flags, total research cost —
    the numbers a reader wants before the table.

Every cell is data the account already carries (scores, QA verdicts, researched
evidence, dossier signals). Nothing is synthesized at export time; empty stays
empty — same honesty contract as the CSV.
"""

from __future__ import annotations

from datetime import UTC, datetime

from openpyxl import Workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from auto_search.scoring.frameworks import FRAMEWORKS

# Quiet, deck-friendly palette (fills need full ARGB).
_HEADER_FILL = PatternFill("solid", fgColor="FF18181B")     # zinc-900
_HEADER_FONT = Font(color="FFFFFFFF", bold=True, size=10)
_BAND_FILL = PatternFill("solid", fgColor="FFF7F7F8")       # zebra rows
_FIT_FILLS = {
    "High": PatternFill("solid", fgColor="FFD1FAE5"),       # emerald-100
    "Medium": PatternFill("solid", fgColor="FFFEF3C7"),     # amber-100
    "Low": PatternFill("solid", fgColor="FFFFEDD5"),        # orange-100
    "Not a fit": PatternFill("solid", fgColor="FFFEE2E2"),  # rose-100
}
_QA_FILLS = {
    "verified": PatternFill("solid", fgColor="FFD1FAE5"),
    "discrepancy": PatternFill("solid", fgColor="FFFEE2E2"),
}
_THIN = Border(bottom=Side(style="thin", color="FFE4E4E7"))
_WRAP = Alignment(wrap_text=True, vertical="top")
_TOP = Alignment(vertical="top")

# (header, width, wrap) — order tells the story: who, verdict, why, receipts.
_COLUMNS = [
    ("Account", 32, False), ("Domain", 22, False), ("Segment", 14, False),
    ("Fit", 10, False), ("Score", 9, False),
    ("Firmographic", 12, False), ("Technographic", 13, False), ("Intent", 9, False),
    ("Recommendation", 70, True), ("QA status", 12, False), ("QA notes", 50, True),
    ("Intent evidence (researched)", 60, True), ("Deep research signals", 60, True),
    ("Discovery signals", 40, True), ("Key facts", 40, True),
    ("Import", 30, False), ("Scored", 11, False), ("Cost (USD)", 10, False),
]


# Stored tier LABELS differ per framework ("Tier 1", "High Fit", …); the BAND
# is the stable vocabulary, and these words match the UI's fitWord (QA F1).
_BAND_WORD = {"high": "High", "medium": "Medium", "low": "Low", "out": "Not a fit"}


def _fit_word(a: dict) -> str:
    return _BAND_WORD.get(a.get("tier_band") or "", a.get("tier_label") or "")


def _pillar_defs(a: dict):
    fw = FRAMEWORKS.get(a.get("framework") or "")
    return fw.pillars if fw else ()


def _pillars(a: dict) -> list[str]:
    """The three board pillars, summed from the framework's OWN dimension
    rollup (health_system folds six dimensions into three) — same math as the
    UI's pillarsFor, never positional (QA F2)."""
    dims = {d.get("key"): d for d in a.get("dimensions") or []}
    out = []
    for p in _pillar_defs(a):
        members = [dims[k] for k in p.dims if k in dims]
        if members:
            score = sum(d.get("score") or 0 for d in members)
            mx = sum(d.get("max") or 0 for d in members)
            out.append(f"{score}/{mx}")
        else:
            out.append("")
    while len(out) < 3:
        out.append("")
    return out[:3]


def _intent_evidence(a: dict) -> str:
    """Summaries of the dimensions in the framework's INTENT pillar — for
    health systems that's competitor + pain + leadership, labelled; for
    specialty/payer it's the single intent dimension (QA F5)."""
    pillar = next((p for p in _pillar_defs(a) if p.key == "intent"), None)
    keys = set(pillar.dims) if pillar else None
    parts = []
    for d in a.get("dimensions") or []:
        key, label = (d.get("key") or ""), (d.get("label") or "")
        member = (key in keys) if keys is not None else ("intent" in (key + label).lower())
        if member and d.get("summary"):
            parts.append(f"{label}: {d['summary']}" if keys and len(keys) > 1 else d["summary"])
    return " | ".join(parts)


def _safe(v):
    """Excel-cell hygiene: strip openpyxl-illegal control chars and neutralize
    formula injection — a '='-leading string (an imported account name, an
    LLM-written summary) must land as text, never as a live formula (QA F3/F6).
    Same contract as the CSV path's csvCell guard."""
    if not isinstance(v, str):
        return v
    v = ILLEGAL_CHARACTERS_RE.sub("", v)
    return "'" + v if v.startswith("=") else v


def _dossier_signals(a: dict) -> str:
    sigs = ((a.get("dossier") or {}).get("intent_signals")) or []
    return "; ".join(
        f"{s.get('signal')}"
        + (f" ({s.get('score')}/10)" if s.get("score") is not None else "")
        + (f": {s.get('detail')}" if s.get("detail") else "")
        for s in sigs)


def _discovery_signals(a: dict) -> str:
    return "; ".join(
        (f"{s.get('signal_type')}: {s.get('summary')}" if s.get("summary") else str(s.get("signal_type")))
        for s in (a.get("discovery_signals") or []))


def _facts(a: dict) -> str:
    return "; ".join(f"{k}: {v}" for k, v in (a.get("firmographics") or {}).items())


def _short_date(iso: str | None) -> str:
    try:
        return datetime.fromisoformat(str(iso).replace("Z", "+00:00")).strftime("%b %d")
    except (ValueError, TypeError):
        return ""


def _row(a: dict) -> list:
    qa = a.get("qa") or {}
    pillars = _pillars(a)
    score = (f"{a.get('total')}/{a.get('max_total')}"
             if a.get("total") is not None and a.get("max_total") is not None else "")
    return [_safe(v) for v in [
        a.get("name"), a.get("domain") or "", (a.get("segment") or "").replace("_", " ").title(),
        _fit_word(a), score,
        pillars[0], pillars[1], pillars[2],
        a.get("recommendation") or "", qa.get("status") or "", qa.get("notes") or "",
        _intent_evidence(a), _dossier_signals(a), _discovery_signals(a), _facts(a),
        a.get("import_label") or "", _short_date(a.get("scored_at")),
        a.get("cost_usd") if a.get("cost_usd") is not None else "",
    ]]


def _style_sheet(ws: Worksheet, n_rows: int) -> None:
    for col, (_header, width, _wrap) in enumerate(_COLUMNS, start=1):
        letter = get_column_letter(col)
        ws.column_dimensions[letter].width = width
        cell = ws.cell(row=1, column=col)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(_COLUMNS))}{n_rows + 1}"
    fit_col = 4
    qa_col = 10
    for r in range(2, n_rows + 2):
        banded = r % 2 == 0
        for c in range(1, len(_COLUMNS) + 1):
            cell = ws.cell(row=r, column=c)
            cell.alignment = _WRAP if _COLUMNS[c - 1][2] else _TOP
            cell.border = _THIN
            if banded and c not in (fit_col, qa_col):
                cell.fill = _BAND_FILL
        fit = ws.cell(row=r, column=fit_col)
        if fit.value in _FIT_FILLS:
            fit.fill = _FIT_FILLS[fit.value]
        qa = ws.cell(row=r, column=qa_col)
        if qa.value in _QA_FILLS:
            qa.fill = _QA_FILLS[qa.value]


def _summary_sheet(ws: Worksheet, accounts: list[dict]) -> None:
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 16
    title = ws.cell(row=1, column=1, value="Scored accounts summary")
    title.font = Font(bold=True, size=13)
    ws.cell(row=2, column=1, value=f"Generated {datetime.now(UTC).strftime('%b %d, %Y %H:%M UTC')}")

    fits: dict[str, int] = {}
    segs: dict[str, int] = {}
    flagged = 0
    cost = 0.0
    for a in accounts:
        fits[_fit_word(a) or "?"] = fits.get(_fit_word(a) or "?", 0) + 1
        seg = (a.get("segment") or "?").replace("_", " ").title()
        segs[seg] = segs.get(seg, 0) + 1
        if (a.get("firmographics") or {}).get("Classification"):
            flagged += 1
        cost += a.get("cost_usd") or 0

    r = 4
    for label, data in (("Fit distribution", fits), ("Segments", segs)):
        head = ws.cell(row=r, column=1, value=label)
        head.font = Font(bold=True)
        r += 1
        order = ["High", "Medium", "Low", "Not a fit"] if label.startswith("Fit") else sorted(data)
        for k in order:
            if k in data:
                ws.cell(row=r, column=1, value=k)
                ws.cell(row=r, column=2, value=data[k])
                r += 1
        r += 1
    head = ws.cell(row=r, column=1, value="Notes")
    head.font = Font(bold=True)
    ws.cell(row=r + 1, column=1, value=f"Accounts: {len(accounts)}")
    ws.cell(row=r + 2, column=1, value=f"Flagged (auto-classified segment): {flagged}")
    ws.cell(row=r + 3, column=1, value=f"Research cost (USD): {round(cost, 2)}")


def build_workbook(accounts: list[dict]) -> Workbook:
    """Accounts (already filtered/ordered by the caller) -> styled workbook."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Scored accounts"
    ws.append([c[0] for c in _COLUMNS])
    for a in accounts:
        ws.append(_row(a))
    _style_sheet(ws, len(accounts))
    _summary_sheet(wb.create_sheet("Summary"), accounts)
    return wb
