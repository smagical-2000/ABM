"""Styled Excel export — tested with the REAL system shapes (the first QA pass
caught tests that fabricated fit labels the platform never produces):
tier_band vocabulary, health_system's six-dimension pillar rollup, formula
injection, and the endpoint's annotate/scored-only/dedupe hygiene."""

import importlib
import io
from datetime import UTC, datetime

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook

from auto_search.scoring import export_xlsx
from auto_search.scoring.models import Account, Dimension, ScoreResult

_app_module = importlib.import_module("auto_search.api.app")


def _hs_account():
    """A health_system account exactly as the platform stores it: six
    dimensions, band 'high', label 'Tier 1' (the framework's real words)."""
    return {
        "account_id": "csv_beaconhealth", "name": "Beacon Health",
        "domain": "beacon.org", "segment": "health_system",
        "framework": "health_system", "state": "scored",
        "tier_band": "high", "tier_label": "Tier 1", "total": 26, "max_total": 30,
        "cost_usd": 0.12, "scored_at": "2026-07-08T17:00:00+00:00",
        "import_label": "SAO cohort", "recommendation": "Lead with prior-auth.",
        "qa": {"status": "verified", "notes": "Revenue confirmed."},
        "dimensions": [
            {"key": "npr", "label": "Net Patient Revenue", "score": 10, "max": 12},
            {"key": "emr", "label": "EMR Compatibility", "score": 4, "max": 5},
            {"key": "competitor", "label": "Competitor Landscape", "score": 3, "max": 4,
             "summary": "UiPath deployed; no AI-RCM vendor."},
            {"key": "pain", "label": "Pain Signals", "score": 3, "max": 4,
             "summary": "RCM hiring wave (Jun 2026)."},
            {"key": "ai_readiness", "label": "Tech Readiness", "score": 1, "max": 2},
            {"key": "leadership", "label": "Leadership", "score": 2, "max": 3,
             "summary": "New CFO (May 2026)."},
        ],
        "dossier": {"intent_signals": [
            {"signal": "RCM hiring wave", "score": 8, "detail": "3 postings in June"},
            {"signal": "No-detail signal", "score": 5}]},
        "discovery_signals": [{"signal_type": "job_posting", "summary": "3 Coder jobs"}],
        "firmographics": {"Net Patient Revenue": "$1.4B"},
    }


def test_workbook_uses_real_bands_pillars_and_evidence():
    not_fit = dict(_hs_account(), account_id="csv_vendor", name="Vendor Co",
                   segment="specialty", framework="specialty",
                   tier_band="out", tier_label="No Fit", total=2,
                   dimensions=[], dossier=None, discovery_signals=[],
                   firmographics={"Classification": "auto-classified specialty (likely not ICP)"},
                   qa={"status": "discrepancy", "notes": "misclassified"})
    wb = export_xlsx.build_workbook([_hs_account(), not_fit])
    ws = wb["Scored accounts"]
    head = [c.value for c in ws[1]]
    row2 = {h: ws.cell(row=2, column=i + 1).value for i, h in enumerate(head)}
    # Fit words come from the BAND (stable across frameworks), not stored labels
    assert row2["Fit"] == "High"
    assert ws.cell(row=3, column=4).value == "Not a fit"
    # health_system pillars are the framework ROLLUP, not dimension positions:
    # techno = emr + ai_readiness, intent = competitor + pain + leadership
    assert row2["Firmographic"] == "10/12"
    assert row2["Technographic"] == "5/7"
    assert row2["Intent"] == "8/11"
    # intent evidence = the intent pillar's labelled summaries
    ev = row2["Intent evidence (researched)"]
    assert "Competitor Landscape: UiPath deployed" in ev
    assert "Pain Signals: RCM hiring wave" in ev and "Leadership: New CFO" in ev
    # dossier detail nit: no trailing colon when detail is missing
    assert row2["Deep research signals"] == (
        "RCM hiring wave (8/10): 3 postings in June; No-detail signal (5/10)")
    # fills fire on the real band words
    assert ws.cell(row=2, column=4).fill.fgColor.rgb == "FFD1FAE5"
    assert ws.cell(row=3, column=4).fill.fgColor.rgb == "FFFEE2E2"
    # summary sheet renders a non-empty fit distribution
    s = wb["Summary"]
    text = " | ".join(str(c.value) for r in s.iter_rows() for c in r if c.value is not None)
    assert "High | 1" in text and "Not a fit | 1" in text
    assert "Flagged (auto-classified segment): 1" in text


def test_cells_are_never_live_formulas_and_control_chars_stripped():
    hostile = dict(_hs_account(), account_id="csv_evil",
                   name='=HYPERLINK("http://evil.example","Q1")',
                   recommendation="=2+2", qa={"status": "verified",
                                              "notes": "bad\x0bchar\x08here"})
    wb = export_xlsx.build_workbook([hostile])
    ws = wb["Scored accounts"]
    name_cell = ws.cell(row=2, column=1)
    rec_cell = ws.cell(row=2, column=9)
    assert name_cell.data_type == "s" and name_cell.value.startswith("'=")
    assert rec_cell.data_type == "s" and rec_cell.value == "'=2+2"
    notes = ws.cell(row=2, column=11).value
    assert notes == "badcharhere"                    # illegal chars stripped


@pytest.fixture
def client(tmp_path, monkeypatch):
    from auto_search.db.engagement_repository import EngagementJsonRepository
    from auto_search.db.repository import JsonFileRepository
    from auto_search.db.scoring_repository import ScoringJsonRepository

    for var in ("BASIC_AUTH_USER", "BASIC_AUTH_PASS", "DATABASE_URL"):
        monkeypatch.delenv(var, raising=False)
    monkeypatch.setattr(_app_module, "get_repository",
                        lambda: JsonFileRepository(tmp_path / "s.json"))
    monkeypatch.setattr(_app_module, "get_scoring_repository",
                        lambda: ScoringJsonRepository(tmp_path / "sc.json"))
    monkeypatch.setattr(_app_module, "get_engagement_repository",
                        lambda: EngagementJsonRepository(tmp_path / "eng.json"))
    with TestClient(_app_module.create_app()) as c:
        yield c


def _seed(repo, name, aid, *, scored=True):
    repo.upsert_account(Account(
        account_id=aid, name=name, segment="specialty",
        framework="specialty", source="csv", domain="x.org"), state="queued")
    if scored:
        repo.save_score(aid, ScoreResult(
            account_id=aid, framework="specialty", framework_version="v",
            dimensions=[Dimension(key="intent", label="Business Priorities & Intent",
                                  score=7, max=10, summary="Hiring wave (Jun 2026).")],
            total=21, max_total=30, tier_band="medium", tier_label="Medium Fit",
            cost_usd=0.1, scored_at=datetime.now(UTC).isoformat()))


def test_endpoint_orders_dedupes_and_exports_scored_only(client):
    repo = client.app.state.scoring_repo
    _seed(repo, "Alpha Health", "csv_alphahealth")
    _seed(repo, "Beta Health", "csv_betahealth")
    _seed(repo, "Parked Co", "csv_parked", scored=False)     # queued: must not export
    ids = ["csv_betahealth", "csv_betahealth", "csv_parked", "csv_alphahealth", 42]
    r = client.post("/api/scoring/export.xlsx", json={"account_ids": ids})
    assert r.status_code == 200
    wb = load_workbook(io.BytesIO(r.content))
    ws = wb["Scored accounts"]
    names = [ws.cell(row=i, column=1).value for i in (2, 3)]
    assert names == ["Beta Health", "Alpha Health"]           # order kept, dupe dropped
    assert ws.cell(row=4, column=1).value is None             # parked row absent
    assert ws.cell(row=2, column=4).value == "Medium"         # band word, annotated row
    # single-dimension frameworks: evidence is the summary, unlabelled
    head = [c.value for c in ws[1]]
    ev_col = head.index("Intent evidence (researched)") + 1
    assert ws.cell(row=2, column=ev_col).value == "Hiring wave (Jun 2026)."
    assert client.post("/api/scoring/export.xlsx", json={}).status_code == 422
    assert client.post("/api/scoring/export.xlsx",
                       json={"account_ids": ["csv_parked"]}).status_code == 404
